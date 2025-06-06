import datetime
from io import BytesIO
from flask import Flask, json, request, jsonify, send_file
from openpyxl import load_workbook
import pandas as pd
import os

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def extract_product_quantity(file_path):
    try:
        df = pd.read_excel(file_path, header=None, skiprows=4)
        if len(df.columns) < 4:
            return {}, 'Excel file does not have enough columns'
        product_col, quantity_col = df.columns[2], df.columns[3]

        df[quantity_col] = df[quantity_col].fillna('')

        filtered_df = df[[product_col, quantity_col]]
        data_dict = {
            str(row[product_col]): row[quantity_col]
            for _, row in filtered_df.iterrows()
            if pd.notnull(row[product_col]) and str(row[product_col]).strip() != ''
        }
        return data_dict, None
    except Exception as e:
        return {}, str(e)

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files.get('file')
    if not file or file.filename == '':
        return jsonify({'error': 'No file provided'}), 400

    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)

    data_dict, error = extract_product_quantity(file_path)
    if error:
        return jsonify({'error': 'Failed to process file', 'message': error}), 500

    return jsonify({'data': data_dict})


@app.route('/export', methods=['POST'])
def update_excel():
    file = request.files.get('file')
    filename = request.form.get('filename')
    json_data = request.form.get('data')

    if not file or file.filename == '':
        return jsonify({'error': 'No Excel file provided'}), 400
    if not filename:
        return jsonify({'error': 'No filename provided'}), 400
    if not json_data:
        return jsonify({'error': 'No JSON data provided'}), 400

    try:
        update_dict = json.loads(json_data)
    except Exception as e:
        return jsonify({'error': 'Invalid JSON format', 'message': str(e)}), 400

    try:
        # Load workbook from uploaded file (in-memory)
        wb = load_workbook(file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=5):
            product_cell = row[2]  # Column C
            quantity_cell = row[3]  # Column D
            product_name = str(product_cell.value).strip() if product_cell.value else ""
            if product_name in update_dict:
                quantity_cell.value = update_dict[product_name]
        
        # Set cell E3 to current date
        sheet['E3'] = datetime.datetime.now().strftime('%d-%m-%Y')
        
        output_stream = BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        return send_file(
            output_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': 'Failed to update Excel file', 'message': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
